"""
Экспорт реестра СОУТ в Excel (XLSX).
Генерирует файл с двумя листами: Направление №1 и Направление №2.
"""
import json
import os
import io
import base64
import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCHEMA = os.environ.get("MAIN_DB_SCHEMA", "t_p19673764_sout_upload_process")

CORS = {
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
    "Access-Control-Allow-Headers": "Content-Type, X-User-Id, X-Auth-Token",
}

COLOR_HEADER_DANGER = "C0392B"
COLOR_HEADER_SAFE = "1A6B3C"
COLOR_SUBHEADER = "1A3050"
COLOR_GOLD = "C8952A"
COLOR_ROW_DANGER = "FDF2F2"
COLOR_ROW_SAFE = "F2FDF5"
COLOR_WHITE = "FFFFFF"


def get_conn():
    return psycopg2.connect(os.environ["DATABASE_URL"])


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def apply_header(ws, row, cols, fill_color):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = thin_border()


def apply_cell(ws, row, col, value, fill_color=None, bold=False, wrap=True, align_h="left"):
    cell = ws.cell(row=row, column=col, value=value)
    if fill_color:
        cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(bold=bold, size=9, name="Calibri")
    cell.alignment = Alignment(horizontal=align_h, vertical="top", wrap_text=wrap)
    cell.border = thin_border()
    return cell


def build_excel(cards_danger, cards_safe, factors_map):
    wb = openpyxl.Workbook()

    # ===== Лист 1: Направление №1 — Опасные факторы =====
    ws1 = wb.active
    ws1.title = "№1 — Опасные факторы"

    # Заголовок листа
    ws1.merge_cells("A1:G1")
    title_cell = ws1["A1"]
    title_cell.value = "НАПРАВЛЕНИЕ №1 — РАБОТНИКИ С ВРЕДНЫМИ (ОПАСНЫМИ) УСЛОВИЯМИ ТРУДА"
    title_cell.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_DANGER)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A2:G2")
    subtitle = ws1["A2"]
    subtitle.value = f"Специальная оценка условий труда · Реестр сформирован системой АВЕСТА · Всего записей: {len(cards_danger)}"
    subtitle.font = Font(size=9, italic=True, color="555555", name="Calibri")
    subtitle.alignment = Alignment(horizontal="center", vertical="center")
    subtitle.fill = PatternFill("solid", fgColor="FCE8E6")
    ws1.row_dimensions[2].height = 18

    cols1 = ["№", "Наименование организации", "Структурное подразделение", "ФИО работника", "Должность (профессия)", "Вредные факторы (код · наименование · расшифровка)", "Дата проведения СОУТ"]
    apply_header(ws1, 3, cols1, COLOR_SUBHEADER)
    ws1.row_dimensions[3].height = 36

    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 28
    ws1.column_dimensions["C"].width = 25
    ws1.column_dimensions["D"].width = 24
    ws1.column_dimensions["E"].width = 22
    ws1.column_dimensions["F"].width = 45
    ws1.column_dimensions["G"].width = 14

    for idx, card in enumerate(cards_danger, 1):
        row = 3 + idx
        factors = factors_map.get(card["id"], [])
        factors_text = ""
        for f in factors:
            desc = f.get("description", "")
            factors_text += f"Класс {f['code']} · {f['name']}"
            if desc:
                factors_text += f": {desc}"
            factors_text += "\n"
        factors_text = factors_text.strip()

        fill = COLOR_ROW_DANGER if idx % 2 == 0 else COLOR_WHITE
        apply_cell(ws1, row, 1, idx, fill, align_h="center")
        apply_cell(ws1, row, 2, card["organization"], fill)
        apply_cell(ws1, row, 3, card["department"], fill)
        apply_cell(ws1, row, 4, card["worker_name"], fill, bold=True)
        apply_cell(ws1, row, 5, card["position"], fill)
        apply_cell(ws1, row, 6, factors_text or "—", fill)
        apply_cell(ws1, row, 7, card["sout_date"] or "—", fill, align_h="center")
        ws1.row_dimensions[row].height = max(30, 15 * len(factors)) if factors else 22

    # ===== Лист 2: Направление №2 — Допустимые условия =====
    ws2 = wb.create_sheet("№2 — Допустимые условия")

    ws2.merge_cells("A1:F1")
    title2 = ws2["A1"]
    title2.value = "НАПРАВЛЕНИЕ №2 — РАБОТНИКИ С ДОПУСТИМЫМИ УСЛОВИЯМИ ТРУДА"
    title2.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    title2.fill = PatternFill("solid", fgColor=COLOR_HEADER_SAFE)
    title2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    ws2.merge_cells("A2:F2")
    sub2 = ws2["A2"]
    sub2.value = f"Специальная оценка условий труда · Реестр сформирован системой АВЕСТА · Всего записей: {len(cards_safe)}"
    sub2.font = Font(size=9, italic=True, color="555555", name="Calibri")
    sub2.alignment = Alignment(horizontal="center", vertical="center")
    sub2.fill = PatternFill("solid", fgColor="EBF9F0")
    ws2.row_dimensions[2].height = 18

    cols2 = ["№", "Наименование организации", "Структурное подразделение", "ФИО работника", "Должность (профессия)", "Дата проведения СОУТ"]
    apply_header(ws2, 3, cols2, COLOR_SUBHEADER)
    ws2.row_dimensions[3].height = 36

    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 26
    ws2.column_dimensions["D"].width = 26
    ws2.column_dimensions["E"].width = 24
    ws2.column_dimensions["F"].width = 16

    for idx, card in enumerate(cards_safe, 1):
        row = 3 + idx
        fill = COLOR_ROW_SAFE if idx % 2 == 0 else COLOR_WHITE
        apply_cell(ws2, row, 1, idx, fill, align_h="center")
        apply_cell(ws2, row, 2, card["organization"], fill)
        apply_cell(ws2, row, 3, card["department"], fill)
        apply_cell(ws2, row, 4, card["worker_name"], fill, bold=True)
        apply_cell(ws2, row, 5, card["position"], fill)
        apply_cell(ws2, row, 6, card["sout_date"] or "—", fill, align_h="center")
        ws2.row_dimensions[row].height = 22

    # ===== Лист 3: Сводка =====
    ws3 = wb.create_sheet("Сводка")
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 20

    ws3.merge_cells("A1:B1")
    t3 = ws3["A1"]
    t3.value = "СВОДНАЯ ИНФОРМАЦИЯ — АВЕСТА"
    t3.font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    t3.fill = PatternFill("solid", fgColor=COLOR_SUBHEADER)
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    summary_data = [
        ("Всего обработано карт СОУТ:", len(cards_danger) + len(cards_safe)),
        ("Направление №1 — с вредными факторами:", len(cards_danger)),
        ("Направление №2 — допустимые условия:", len(cards_safe)),
        ("Доля вредных условий (%):", round(len(cards_danger) / max(len(cards_danger) + len(cards_safe), 1) * 100, 1)),
    ]
    for i, (label, value) in enumerate(summary_data, 2):
        apply_cell(ws3, i, 1, label, bold=True)
        apply_cell(ws3, i, 2, value, align_h="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def handler(event: dict, context) -> dict:
    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": CORS, "body": ""}

    params = event.get("queryStringParameters") or {}
    batch_id = params.get("batch_id")
    direction = params.get("direction", "all")

    conn = get_conn()
    cur = conn.cursor()

    where = ""
    args = []
    if batch_id:
        where = "WHERE c.batch_id = %s"
        args = [batch_id]

    cur.execute(f"""
        SELECT c.id, c.organization, c.department, c.worker_name, c.position, c.sout_date, c.is_dangerous
        FROM {SCHEMA}.sout_cards c {where}
        ORDER BY c.is_dangerous DESC, c.organization, c.worker_name
    """, args)
    rows = cur.fetchall()

    all_cards = []
    for r in rows:
        all_cards.append({
            "id": r[0], "organization": r[1] or "", "department": r[2] or "",
            "worker_name": r[3] or "", "position": r[4] or "",
            "sout_date": r[5] or "", "is_dangerous": r[6],
        })

    card_ids = [c["id"] for c in all_cards]
    factors_map: dict = {}
    if card_ids:
        placeholders = ",".join(["%s"] * len(card_ids))
        cur.execute(
            f"SELECT card_id, code, name, description FROM {SCHEMA}.sout_factors WHERE card_id IN ({placeholders})",
            card_ids,
        )
        for fr in cur.fetchall():
            cid = fr[0]
            if cid not in factors_map:
                factors_map[cid] = []
            factors_map[cid].append({"code": fr[1], "name": fr[2], "description": fr[3] or ""})

    cur.close()
    conn.close()

    if direction == "danger":
        cards_danger = [c for c in all_cards if c["is_dangerous"]]
        cards_safe = []
    elif direction == "safe":
        cards_danger = []
        cards_safe = [c for c in all_cards if not c["is_dangerous"]]
    else:
        cards_danger = [c for c in all_cards if c["is_dangerous"]]
        cards_safe = [c for c in all_cards if not c["is_dangerous"]]

    xlsx_bytes = build_excel(cards_danger, cards_safe, factors_map)
    b64 = base64.b64encode(xlsx_bytes).decode("utf-8")

    return {
        "statusCode": 200,
        "headers": {
            **CORS,
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": "attachment; filename=AVESTA_SOUT_Reestr.xlsx",
        },
        "body": b64,
        "isBase64Encoded": True,
    }
